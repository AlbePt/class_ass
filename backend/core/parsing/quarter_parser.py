from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.core.models import ParsedEntry, ParsedWorkbook, StudentSection

MONTH_ALIASES = {
    "январь": 1,
    "янв": 1,
    "февраль": 2,
    "фев": 2,
    "март": 3,
    "мар": 3,
    "апрель": 4,
    "апр": 4,
    "май": 5,
    "июнь": 6,
    "июл": 7,
    "июль": 7,
    "август": 8,
    "авг": 8,
    "сентябрь": 9,
    "сен": 9,
    "октябрь": 10,
    "окт": 10,
    "ноябрь": 11,
    "ноя": 11,
    "декабрь": 12,
    "дек": 12,
}

ATTENDANCE_DEFAULT = {"Н", "У", "Б", "О"}
META_PREFIXES = {"школа", "учебный год", "класс", "период"}
TOKEN_SPLIT_RE = re.compile(r"[\s,;]+")


@dataclass
class SectionParseResult:
    section: StudentSection
    end_row: int
    updated_academic_year_start: Optional[int]
    updated_academic_year_end: Optional[int]


class QuarterReportParser:
    """Parser for quarterly performance reports according to v2 specification."""

    def parse_workbook(self, xlsx_bytes: bytes) -> ParsedWorkbook:
        wb = load_workbook(filename=BytesIO(xlsx_bytes), data_only=True)
        sheet = self._find_target_sheet(wb.worksheets)
        if sheet is None:
            raise ValueError("Не удалось найти лист с данными (отсутствует строка 'Предмет').")

        grid = self._expand_grid(sheet)
        max_row = len(grid) - 1
        workbook_meta = {
            "school_name": None,
            "academic_year_start": None,
            "academic_year_end": None,
        }
        students: List[StudentSection] = []
        global_warnings: List[str] = []

        row = 1
        while row <= max_row:
            cell_value = grid[row][1]
            if cell_value:
                normalized = self._normalize_header(cell_value)
                if normalized.startswith("школа"):
                    workbook_meta["school_name"] = self._extract_value(cell_value)
                elif normalized.startswith("учебный год"):
                    year_pair = self._parse_academic_year(cell_value)
                    if year_pair:
                        workbook_meta["academic_year_start"], workbook_meta["academic_year_end"] = year_pair
                elif normalized.startswith("класс"):
                    # class meta maintained, but handled inside sections
                    pass
                elif normalized.startswith("период"):
                    pass
                elif normalized.startswith("ученик"):
                    result = self._parse_student_section(
                        grid,
                        start_row=row,
                        base_meta=workbook_meta,
                        global_warnings=global_warnings,
                    )
                    students.append(result.section)
                    if result.updated_academic_year_start is not None:
                        workbook_meta["academic_year_start"] = result.updated_academic_year_start
                    if result.updated_academic_year_end is not None:
                        workbook_meta["academic_year_end"] = result.updated_academic_year_end
                    row = result.end_row
                # fall-through to increment row below
            row += 1

        return ParsedWorkbook(
            school_name=workbook_meta.get("school_name"),
            academic_year_start=workbook_meta.get("academic_year_start"),
            academic_year_end=workbook_meta.get("academic_year_end"),
            students=students,
            global_warnings=global_warnings,
        )

    # ------------------------------------------------------------------
    # Sheet preparation helpers
    # ------------------------------------------------------------------
    def _find_target_sheet(self, sheets: List[Worksheet]) -> Optional[Worksheet]:
        for sheet in sheets:
            for row in sheet.iter_rows(values_only=True):
                if row and any(self._normalize_header(str(cell)) == "предмет" for cell in row if cell is not None):
                    return sheet
        return None

    def _expand_grid(self, sheet: Worksheet) -> List[List[Optional[str]]]:
        max_row = sheet.max_row
        max_col = sheet.max_column
        grid: List[List[Optional[str]]] = [[None for _ in range(max_col + 1)] for _ in range(max_row + 1)]
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                value = sheet.cell(row=row, column=col).value
                grid[row][col] = self._normalize_value(value)
        for merged in sheet.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
            value = grid[min_row][min_col]
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    grid[row][col] = value
        return grid

    # ------------------------------------------------------------------
    # Parsing helpers
    # ------------------------------------------------------------------
    def _parse_student_section(
        self,
        grid: List[List[Optional[str]]],
        start_row: int,
        base_meta: Dict[str, Optional[str]],
        global_warnings: List[str],
    ) -> SectionParseResult:
        max_row = len(grid) - 1
        header = grid[start_row][1] or ""
        fio_raw = self._extract_value(header)
        fio_norm = self._normalize_name(fio_raw)

        klass = ""
        period_from: Optional[date] = None
        period_to: Optional[date] = None
        academic_year_start = base_meta.get("academic_year_start")
        academic_year_end = base_meta.get("academic_year_end")

        row = start_row + 1
        row_header_months = -1
        row_days = -1
        attendance_legend: Dict[str, str] = {}
        entries: List[ParsedEntry] = []
        warnings: List[str] = []

        while row <= max_row:
            first_col = grid[row][1]
            normalized = self._normalize_header(first_col) if first_col else ""
            if not first_col:
                row += 1
                continue

            if normalized.startswith("ученик") and row != start_row:
                row -= 1
                break
            if normalized.startswith("учебный год"):
                year_pair = self._parse_academic_year(first_col)
                if year_pair:
                    academic_year_start, academic_year_end = year_pair
                row += 1
                continue
            if normalized.startswith("класс"):
                klass = self._extract_value(first_col)
                row += 1
                continue
            if normalized.startswith("период"):
                p = self._parse_period(first_col)
                if p:
                    period_from, period_to = p
                row += 1
                continue
            if normalized.startswith("школа"):
                row += 1
                continue
            if normalized == "предмет":
                row_header_months = row
                row_days = row + 1
                break
            row += 1

        if row_header_months == -1 or row_days == -1 or row_days > max_row:
            warnings.append("Не найдена таблица предметов для ученика")
            section = StudentSection(
                fio_raw=fio_raw,
                fio_norm=fio_norm,
                klass=klass,
                period_from=period_from,
                period_to=period_to,
                entries=entries,
                attendance_legend=attendance_legend,
                warnings=warnings,
            )
            return SectionParseResult(section, row, academic_year_start, academic_year_end)

        col_date_map, mapping_warnings = self._build_date_mapping(
            grid,
            row_header_months,
            row_days,
            academic_year_start,
            academic_year_end,
            period_from,
            period_to,
        )
        warnings.extend(mapping_warnings)

        attendance_codes = set(ATTENDANCE_DEFAULT)
        current_row = row_days + 1
        while current_row <= max_row:
            first_col = grid[current_row][1]
            normalized = self._normalize_header(first_col) if first_col else ""

            if self._is_legend_row(grid, current_row):
                code = (grid[current_row][1] or "").strip()
                description = (grid[current_row][2] or "").strip()
                if code:
                    attendance_legend[code] = description
                    attendance_codes.add(code)
                current_row += 1
                continue

            if not first_col:
                # potential end of table, but continue scanning for legend/new sections
                next_nonempty = self._find_next_nonempty(grid, current_row + 1)
                if next_nonempty and self._normalize_header(next_nonempty).startswith("ученик"):
                    break
                current_row += 1
                continue

            if normalized in META_PREFIXES or normalized.startswith("ученик"):
                current_row -= 1
                break

            subject_name = first_col.strip()
            if not subject_name:
                current_row += 1
                continue

            subject_entries_found = False
            for col, mapped_date in col_date_map.items():
                raw = grid[current_row][col]
                if raw is None or str(raw).strip() == "":
                    continue
                raw_text = str(raw)
                grades, attendance, token_warnings, multi_flag = self._tokenize_cell(raw_text, attendance_codes)
                for warning in token_warnings:
                    warnings.append(
                        f"[{fio_norm}] {subject_name}: {warning} (row={current_row}, col={col})"
                    )
                if multi_flag:
                    warnings.append(
                        f"[{fio_norm}] {subject_name}: multiple_tokens_in_cell (row={current_row}, col={col})"
                    )
                entry = ParsedEntry(
                    student_fio_raw=fio_raw,
                    student_fio_norm=fio_norm,
                    klass=klass,
                    subject=subject_name,
                    date=mapped_date,
                    grades=grades,
                    attendance=attendance,
                    raw_text=raw_text,
                    row=current_row,
                    col=col,
                )
                if period_from and period_to and not (period_from <= mapped_date <= period_to):
                    warnings.append(
                        f"[{fio_norm}] {subject_name}: date_out_of_period {mapped_date.isoformat()}"
                    )
                entries.append(entry)
                subject_entries_found = True
            if not subject_entries_found:
                # still allow as subject with no entries
                pass
            current_row += 1

        section = StudentSection(
            fio_raw=fio_raw,
            fio_norm=fio_norm,
            klass=klass,
            period_from=period_from,
            period_to=period_to,
            entries=entries,
            attendance_legend=attendance_legend,
            warnings=warnings,
        )
        return SectionParseResult(section, current_row, academic_year_start, academic_year_end)

    def _build_date_mapping(
        self,
        grid: List[List[Optional[str]]],
        row_months: int,
        row_days: int,
        academic_year_start: Optional[int],
        academic_year_end: Optional[int],
        period_from: Optional[date],
        period_to: Optional[date],
    ) -> Tuple[Dict[int, date], List[str]]:
        warnings: List[str] = []
        col_date_map: Dict[int, date] = {}
        max_col = len(grid[row_months]) - 1
        current_month: Optional[int] = None
        for col in range(2, max_col + 1):
            header_value = grid[row_months][col]
            if header_value:
                month_number = self._parse_month(header_value)
                if month_number:
                    current_month = month_number
                else:
                    warnings.append(f"Неизвестный месяц: {header_value}")
                    current_month = None
            if current_month is None:
                continue
            day_raw = grid[row_days][col]
            if not day_raw:
                continue
            try:
                day = int(str(day_raw).strip())
            except ValueError:
                warnings.append(f"Некорректный день: {day_raw}")
                continue
            year = self._resolve_year(current_month, day, academic_year_start, academic_year_end, period_from, period_to)
            if year is None:
                warnings.append(f"Не удалось определить год для {current_month}.{day}")
                continue
            try:
                mapped_date = date(year, current_month, day)
            except ValueError:
                warnings.append(f"Некорректная дата: {current_month}.{day}.{year}")
                continue
            col_date_map[col] = mapped_date
        return col_date_map, warnings

    def _resolve_year(
        self,
        month: int,
        day: int,
        academic_year_start: Optional[int],
        academic_year_end: Optional[int],
        period_from: Optional[date],
        period_to: Optional[date],
    ) -> Optional[int]:
        if academic_year_start and academic_year_end:
            if month >= 9:
                return academic_year_start
            return academic_year_end
        candidate_years = []
        if period_from:
            candidate_years.append(period_from.year)
        if period_to and period_to.year not in candidate_years:
            candidate_years.append(period_to.year)
        if not candidate_years:
            candidate_years.append(datetime.utcnow().year)
        for year in candidate_years:
            try:
                dt = date(year, month, day)
            except ValueError:
                continue
            if period_from and period_to:
                if period_from <= dt <= period_to:
                    return year
            else:
                return year
        if period_from and period_to:
            # fallback cross-year guess
            if month >= 9:
                return period_from.year
            return period_to.year
        return None

    def _tokenize_cell(
        self, raw_text: str, attendance_codes: set[str]
    ) -> Tuple[List[int], List[str], List[str], bool]:
        cleaned = raw_text.replace("/", " ")
        tokens = [tok for tok in TOKEN_SPLIT_RE.split(cleaned) if tok]
        grades: List[int] = []
        attendance: List[str] = []
        warnings: List[str] = []
        multiple = len(tokens) > 1
        for token in tokens:
            token = token.strip()
            if not token:
                continue
            if token in attendance_codes:
                attendance.append(token)
                continue
            if token.isdigit():
                try:
                    value = int(token)
                except ValueError:
                    warnings.append(f"unknown_token {token}")
                    continue
                if 2 <= value <= 5:
                    grades.append(value)
                else:
                    warnings.append(f"invalid_grade {token}")
                continue
            warnings.append(f"unknown_token {token}")
        return grades, attendance, warnings, multiple

    def _parse_month(self, value: Optional[str]) -> Optional[int]:
        if not value:
            return None
        key = str(value).strip().lower()
        key = key.replace("\xa0", " ")
        key = " ".join(key.split())
        return MONTH_ALIASES.get(key)

    def _parse_academic_year(self, value: Optional[str]) -> Optional[Tuple[int, int]]:
        if not value:
            return None
        text = str(value)
        match = re.search(r"(\d{4})\s*[/\\-]\s*(\d{4})", text)
        if match:
            start = int(match.group(1))
            end = int(match.group(2))
            if end < start:
                end = start + 1
            return start, end
        return None

    def _parse_period(self, value: Optional[str]) -> Optional[Tuple[date, date]]:
        if not value:
            return None
        text = str(value)
        match = re.search(r"с\s*(\d{1,2}[./]\d{1,2}[./]\d{2,4}).*по\s*(\d{1,2}[./]\d{1,2}[./]\d{2,4})", text, re.IGNORECASE)
        if not match:
            return None
        start = self._parse_date(match.group(1))
        end = self._parse_date(match.group(2))
        if start and end:
            return start, end
        return None

    def _parse_date(self, raw: str) -> Optional[date]:
        normalized = raw.replace("/", ".").replace(" ", "")
        parts = normalized.split(".")
        if len(parts) != 3:
            return None
        day, month, year = parts
        if len(year) == 2:
            year = "20" + year
        try:
            return date(int(year), int(month), int(day))
        except ValueError:
            return None

    def _extract_value(self, text: Optional[str]) -> str:
        if text is None:
            return ""
        raw = str(text).replace("\xa0", " ")
        parts = raw.split(":", 1)
        if len(parts) == 2:
            return parts[1].strip()
        return raw.strip()

    def _normalize_value(self, value: Optional[object]) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, str):
            cleaned = value.replace("\xa0", " ")
            cleaned = cleaned.strip()
            return cleaned
        return str(value).strip()

    def _normalize_header(self, value: Optional[str]) -> str:
        if value is None:
            return ""
        cleaned = self._normalize_value(value) or ""
        return cleaned.lower()

    def _normalize_name(self, value: str) -> str:
        return " ".join(value.split())

    def _is_legend_row(self, grid: List[List[Optional[str]]], row: int) -> bool:
        code = grid[row][1] or ""
        description = grid[row][2] or ""
        if not code or not description:
            return False
        code = code.strip()
        description = description.strip()
        if len(code) != 1:
            return False
        if any(grid[row][col] for col in range(3, len(grid[row]))):
            return False
        return True

    def _find_next_nonempty(self, grid: List[List[Optional[str]]], start_row: int) -> Optional[str]:
        max_row = len(grid) - 1
        for row in range(start_row, max_row + 1):
            value = grid[row][1]
            if value:
                return value
        return None


__all__ = ["QuarterReportParser"]
