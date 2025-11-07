from io import BytesIO

from openpyxl import Workbook

from backend.core.parsing.quarter_parser import QuarterReportParser


def build_workbook_with_legend() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Учебный год: 2024/2025'
    ws['A2'] = 'Период: с 01.09.2024 по 30.09.2024'
    ws['A3'] = 'Ученик: Петров Пётр'
    ws['A5'] = 'Предмет'
    ws['B5'] = 'Сентябрь'
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=3)
    ws['B6'] = 1
    ws['C6'] = 2
    ws['A7'] = 'История'
    ws['B7'] = '4'
    ws['C7'] = 'Н'
    ws['A9'] = 'Н'
    ws['B9'] = 'Неуважительная'
    ws['A10'] = 'У'
    ws['B10'] = 'Уважительная'

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_legend_detected_and_attendance_collected():
    parser = QuarterReportParser()
    workbook = parser.parse_workbook(build_workbook_with_legend())
    section = workbook.students[0]
    assert section.attendance_legend == {'Н': 'Неуважительная', 'У': 'Уважительная'}
    history_entries = [e for e in section.entries if e.subject == 'История']
    assert any('Н' in entry.attendance for entry in history_entries)
