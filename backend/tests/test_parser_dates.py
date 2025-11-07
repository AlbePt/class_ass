from datetime import date
from io import BytesIO

from openpyxl import Workbook

from backend.core.parsing.quarter_parser import QuarterReportParser


def build_sample_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    ws['A1'] = 'Учебный год: 2025/2026'
    ws['A2'] = 'Класс: 9А'
    ws['A3'] = 'Период: с 01.09.2025 по 24.10.2025'
    ws['A4'] = 'Ученик: Иванов Иван'
    ws['A6'] = 'Предмет'
    ws['B6'] = 'Сентябрь'
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=3)
    ws['B7'] = 1
    ws['C7'] = 2
    ws['A8'] = 'Математика'
    ws['B8'] = '5'
    ws['C8'] = '4'
    ws['A9'] = 'Русский язык'
    ws['B9'] = '3'
    ws['C9'] = '2'

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_dates_resolved_with_academic_year():
    data = build_sample_workbook()
    parser = QuarterReportParser()
    workbook = parser.parse_workbook(data)
    assert len(workbook.students) == 1
    student = workbook.students[0]
    dates = sorted({entry.date for entry in student.entries})
    assert dates == [date(2025, 9, 1), date(2025, 9, 2)]
