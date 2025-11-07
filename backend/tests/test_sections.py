from io import BytesIO

from openpyxl import Workbook

from backend.core.parsing.quarter_parser import QuarterReportParser


def build_multi_student_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Учебный год: 2025/2026'
    ws['A2'] = 'Период: с 01.09.2025 по 30.09.2025'
    ws['A3'] = 'Ученик: Иванов И.И.'
    ws['A5'] = 'Предмет'
    ws['B5'] = 'Сентябрь'
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=3)
    ws['B6'] = 1
    ws['C6'] = 2
    ws['A7'] = 'Математика'
    ws['B7'] = '5'
    ws['C7'] = '5'
    ws['A9'] = 'Ученик: Сидорова А.А.'
    ws['A11'] = 'Предмет'
    ws['B11'] = 'Сентябрь'
    ws.merge_cells(start_row=11, start_column=2, end_row=11, end_column=3)
    ws['B12'] = 1
    ws['C12'] = 2
    ws['A13'] = 'История'
    ws['B13'] = '4'
    ws['C13'] = '5'

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_multiple_students_parsed():
    parser = QuarterReportParser()
    workbook = parser.parse_workbook(build_multi_student_workbook())
    assert len(workbook.students) == 2
    assert {student.fio_norm for student in workbook.students} == {
        'Иванов И.И.',
        'Сидорова А.А.',
    }
